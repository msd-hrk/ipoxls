import openpyxl.worksheet
from openpyxl.styles import Alignment

class BindingDesigner():
    
    def __init__(self) -> None:
        self.align_cent = Alignment(horizontal='center')
        self.align_left = Alignment(horizontal='left')
        self.align_right = Alignment(horizontal='right')
        self.align_just_top = Alignment(horizontal='justify', vertical='top')
        self.align_cc = Alignment(horizontal='center', vertical='center')
        self.align_rc = Alignment(horizontal='right', vertical='center')
        pass
    
    def create_individual_binding(self, sheet: openpyxl.worksheet):
        self.company_summary(sheet)
        self.ipo_info(sheet)
        self.listing_forecast(sheet)
        self.financial_info(sheet)
        self.news_list(sheet)
        self.lockup_info(sheet)
        pass
    
    def company_summary(self, sheet: openpyxl.worksheet):
        # merge cells
        sheet.merge_cells('D1:O1')
        sheet.merge_cells('B3:O3')
        sheet.merge_cells('A4:A15')
        sheet.merge_cells('B4:O15')
        sheet.merge_cells('A17:G17')
        sheet.merge_cells('B18:C18')
        sheet.merge_cells('D18:D20')
        sheet.merge_cells('B19:C19')
        sheet.merge_cells('B20:C20')
        sheet.merge_cells('F18:G18')
        sheet.merge_cells('F19:G19')
        sheet.merge_cells('F20:G20')
        
        sheet['A1'].alignment = self.align_cent
        sheet['B1'].alignment = self.align_cent
        sheet['C1'].alignment = self.align_cent
        sheet['D1'].alignment = self.align_cent
        sheet['A3'].alignment = self.align_cent
        sheet['A4'].alignment = self.align_cc
        sheet['B4'].alignment = self.align_just_top
        sheet['A17'].alignment = self.align_cent
        sheet['A18'].alignment = self.align_cent
        sheet['B18'].alignment = self.align_cent
        sheet['D18'].alignment = self.align_cc
        sheet['E18'].alignment = self.align_cent
        sheet['F18'].alignment = self.align_cent
        sheet['A19'].alignment = self.align_cent
        sheet['B19'].alignment = self.align_cent
        sheet['E19'].alignment = self.align_cent
        sheet['F19'].alignment = self.align_cent
        sheet['A20'].alignment = self.align_cent
        sheet['B20'].alignment = self.align_cent
        sheet['E20'].alignment = self.align_cent
        sheet['F20'].alignment = self.align_cent
        pass
    
    def ipo_info(self, sheet: openpyxl.worksheet):
        sheet.merge_cells('A22:G22')
        sheet.merge_cells('A23:B23')
        sheet.merge_cells('C23:D23')
        sheet.merge_cells('F23:G23')
        sheet.merge_cells('A24:B24')
        sheet.merge_cells('C24:D24')
        sheet.merge_cells('F24:G24')
        sheet.merge_cells('A25:B25')
        sheet.merge_cells('C25:D25')
        sheet.merge_cells('F25:G25')
        sheet.merge_cells('B26:C26')
        sheet.merge_cells('D26:E26')
        sheet.merge_cells('C27:C29')
        sheet.merge_cells('D27:E29')
        sheet.merge_cells('A30:B30')
        sheet.merge_cells('E30:F30')
        sheet.merge_cells('A31:B31')
        sheet.merge_cells('E31:F31')
        
        sheet['A22'].alignment = self.align_cent
        sheet['A23'].alignment = self.align_cent
        sheet['C23'].alignment = self.align_cent
        sheet['E23'].alignment = self.align_cent
        sheet['F23'].alignment = self.align_cent
        sheet['A24'].alignment = self.align_cent
        sheet['C24'].alignment = self.align_cent
        sheet['E24'].alignment = self.align_cent
        sheet['F24'].alignment = self.align_cent
        sheet['A25'].alignment = self.align_cent
        sheet['C25'].alignment = self.align_cent
        sheet['E25'].alignment = self.align_cent
        sheet['F25'].alignment = self.align_cent
        sheet['A26'].alignment = self.align_cent
        sheet['B26'].alignment = self.align_cent
        sheet['D26'].alignment = self.align_cent
        sheet['F26'].alignment = self.align_cent
        sheet['G26'].alignment = self.align_cent
        sheet['A27'].alignment = self.align_cent
        sheet['B27'].alignment = self.align_right
        sheet['C27'].alignment = self.align_rc
        sheet['D27'].alignment = self.align_cc
        sheet['F27'].alignment = self.align_cent
        sheet['G27'].alignment = self.align_cent
        sheet['A28'].alignment = self.align_cent
        sheet['B28'].alignment = self.align_right
        sheet['F28'].alignment = self.align_cent
        sheet['G28'].alignment = self.align_cent
        sheet['A29'].alignment = self.align_cent
        sheet['B29'].alignment = self.align_right
        sheet['F29'].alignment = self.align_cent
        sheet['G29'].alignment = self.align_cent
        sheet['A30'].alignment = self.align_cent
        sheet['C30'].alignment = self.align_right
        sheet['D30'].alignment = self.align_cent
        sheet['E30'].alignment = self.align_cent
        sheet['G30'].alignment = self.align_cent
        sheet['A31'].alignment = self.align_cent
        sheet['C31'].alignment = self.align_cent
        sheet['D31'].alignment = self.align_cent
        sheet['E31'].alignment = self.align_cent
        sheet['G31'].alignment = self.align_cent
        pass
    
    def listing_forecast(self, sheet: openpyxl.worksheet):
        sheet.merge_cells('A33:G33')
        sheet.merge_cells('A34:D34')
        sheet.merge_cells('E34:G34')
        sheet.merge_cells('B35:D35')
        sheet.merge_cells('F35:G35')
        sheet.merge_cells('B36:D36')
        sheet.merge_cells('F36:G36')
        sheet.merge_cells('B37:D37')
        sheet.merge_cells('F37:G37')
        sheet.merge_cells('B38:D38')
        sheet.merge_cells('F38:G38')
        sheet.merge_cells('B39:D39')

        sheet['A33'].alignment = self.align_cent
        sheet['A34'].alignment = self.align_cent
        sheet['E34'].alignment = self.align_cent
        sheet['A35'].alignment = self.align_cent
        sheet['B35'].alignment = self.align_cent
        sheet['E35'].alignment = self.align_cent
        sheet['F35'].alignment = self.align_cent
        sheet['A36'].alignment = self.align_cent
        sheet['B36'].alignment = self.align_cent
        sheet['E36'].alignment = self.align_cent
        sheet['F36'].alignment = self.align_cent
        sheet['A37'].alignment = self.align_cent
        sheet['B37'].alignment = self.align_cent
        sheet['E37'].alignment = self.align_cent
        sheet['F37'].alignment = self.align_cent
        sheet['A38'].alignment = self.align_cent
        sheet['B38'].alignment = self.align_cent
        sheet['E38'].alignment = self.align_cent
        sheet['F38'].alignment = self.align_cent
        sheet['A39'].alignment = self.align_cent
        sheet['B39'].alignment = self.align_cent
        pass
    
    def financial_info(self, sheet: openpyxl.worksheet):
        # independent financial infomation
        sheet.merge_cells('A41:G41')
        sheet.merge_cells('B42:C42')
        sheet.merge_cells('D42:E42')
        sheet.merge_cells('F42:G42')
        sheet.merge_cells('B43:C43')
        sheet.merge_cells('D43:E43')
        sheet.merge_cells('F43:G43')
        sheet.merge_cells('B44:C44')
        sheet.merge_cells('D44:E44')
        sheet.merge_cells('F44:G44')
        sheet.merge_cells('B45:C45')
        sheet.merge_cells('D45:E45')
        sheet.merge_cells('F45:G45')
        sheet.merge_cells('B46:C46')
        sheet.merge_cells('D46:E46')
        sheet.merge_cells('F46:G46')
        sheet.merge_cells('B47:C47')
        sheet.merge_cells('D47:E47')
        sheet.merge_cells('F47:G47')
        sheet.merge_cells('B48:C48')
        sheet.merge_cells('D48:E48')
        sheet.merge_cells('F48:G48')
        sheet.merge_cells('B49:C49')
        sheet.merge_cells('D49:E49')
        sheet.merge_cells('F49:G49')
        
        sheet['A41'].alignment = self.align_cent
        sheet['A42'].alignment = self.align_cent
        sheet['A43'].alignment = self.align_cent
        sheet['A44'].alignment = self.align_cent
        sheet['A45'].alignment = self.align_cent
        sheet['A46'].alignment = self.align_cent
        sheet['A47'].alignment = self.align_cent
        sheet['A48'].alignment = self.align_cent
        sheet['A49'].alignment = self.align_cent
        sheet['B42'].alignment = self.align_cent
        sheet['B43'].alignment = self.align_cent
        sheet['B44'].alignment = self.align_cent
        sheet['B45'].alignment = self.align_cent
        sheet['B46'].alignment = self.align_cent
        sheet['B47'].alignment = self.align_cent
        sheet['B48'].alignment = self.align_cent
        sheet['B49'].alignment = self.align_cent
        sheet['D42'].alignment = self.align_cent
        sheet['D43'].alignment = self.align_cent
        sheet['D44'].alignment = self.align_cent
        sheet['D45'].alignment = self.align_cent
        sheet['D46'].alignment = self.align_cent
        sheet['D47'].alignment = self.align_cent
        sheet['D48'].alignment = self.align_cent
        sheet['D49'].alignment = self.align_cent
        sheet['F42'].alignment = self.align_cent
        sheet['F43'].alignment = self.align_cent
        sheet['F44'].alignment = self.align_cent
        sheet['F45'].alignment = self.align_cent
        sheet['F46'].alignment = self.align_cent
        sheet['F47'].alignment = self.align_cent
        sheet['F48'].alignment = self.align_cent
        sheet['F49'].alignment = self.align_cent

        # consolidated financial information
        sheet.merge_cells('A51:G51')
        sheet.merge_cells('B52:C52')
        sheet.merge_cells('D52:E52')
        sheet.merge_cells('F52:G52')
        sheet.merge_cells('B53:C53')
        sheet.merge_cells('D53:E53')
        sheet.merge_cells('F53:G53')
        sheet.merge_cells('B54:C54')
        sheet.merge_cells('D54:E54')
        sheet.merge_cells('F54:G54')
        sheet.merge_cells('B55:C55')
        sheet.merge_cells('D55:E55')
        sheet.merge_cells('F55:G55')
        sheet.merge_cells('B56:C56')
        sheet.merge_cells('D56:E56')
        sheet.merge_cells('F56:G56')
        sheet.merge_cells('B57:C57')
        sheet.merge_cells('D57:E57')
        sheet.merge_cells('F57:G57')
        sheet.merge_cells('B58:C58')
        sheet.merge_cells('D58:E58')
        sheet.merge_cells('F58:G58')
        
        sheet['A51'].alignment = self.align_cent
        sheet['A52'].alignment = self.align_cent
        sheet['A53'].alignment = self.align_cent
        sheet['A54'].alignment = self.align_cent
        sheet['A55'].alignment = self.align_cent
        sheet['A56'].alignment = self.align_cent
        sheet['A57'].alignment = self.align_cent
        sheet['A58'].alignment = self.align_cent
        sheet['B52'].alignment = self.align_cent
        sheet['B53'].alignment = self.align_cent
        sheet['B54'].alignment = self.align_cent
        sheet['B55'].alignment = self.align_cent
        sheet['B56'].alignment = self.align_cent
        sheet['B57'].alignment = self.align_cent
        sheet['B58'].alignment = self.align_cent
        sheet['D52'].alignment = self.align_cent
        sheet['D53'].alignment = self.align_cent
        sheet['D54'].alignment = self.align_cent
        sheet['D55'].alignment = self.align_cent
        sheet['D56'].alignment = self.align_cent
        sheet['D57'].alignment = self.align_cent
        sheet['D58'].alignment = self.align_cent
        sheet['F52'].alignment = self.align_cent
        sheet['F53'].alignment = self.align_cent
        sheet['F54'].alignment = self.align_cent
        sheet['F55'].alignment = self.align_cent
        sheet['F56'].alignment = self.align_cent
        sheet['F57'].alignment = self.align_cent
        sheet['F58'].alignment = self.align_cent

    def news_list(self, sheet: openpyxl.worksheet):
        cell_row = 60
        tmp = 'A' + str(cell_row) + ':G' + str(cell_row)
        sheet.merge_cells(tmp)
        sheet['A' + str(cell_row)].alignment = self.align_cent
        
        tmp = 'A' + str(cell_row + 1) + ':B' + str(cell_row + 1)
        sheet.merge_cells(tmp)
        sheet['A' + str(cell_row + 1)].alignment = self.align_cent
        
        tmp = 'C' + str(cell_row + 1) + ':G' + str(cell_row + 1)
        sheet.merge_cells(tmp)
        sheet['C' + str(cell_row + 1)].alignment = self.align_cent
        
        for i in range(cell_row + 2, cell_row + 21):
            tmp = 'A' + str(i) + ':B' + str(i)
            sheet.merge_cells(tmp)
            sheet['A' + str(i)].alignment = self.align_cent
            tmp = 'C' + str(i) + ':G' + str(i)
            sheet.merge_cells(tmp)
            sheet['C' + str(i)].alignment = self.align_left
            

    def lockup_info(self, sheet: openpyxl.worksheet):
        cell_row = 82
        sheet.merge_cells('A' + str(cell_row) + ':G' + str(cell_row))
        sheet['A' + str(cell_row)].alignment = self.align_cent
        
        for row in range(cell_row + 1, cell_row + 40, 4):
            sheet.merge_cells(start_row=row, start_column=1, end_row=row+3, end_column=1)
            sheet.merge_cells(start_row=row, start_column=4, end_row=row, end_column=7)
            sheet.merge_cells(start_row=row+1, start_column=2, end_row=row+2, end_column=2)
            sheet.merge_cells(start_row=row+1, start_column=4, end_row=row+1, end_column=7)
            
            sheet["A" + str(row)].alignment = self.align_cc
            sheet["B" + str(row)].alignment = self.align_cent
            sheet["C" + str(row)].alignment = self.align_cent
            sheet["D" + str(row)].alignment = self.align_cent
            sheet["B" + str(row + 1)].alignment = self.align_cc
            sheet["C" + str(row + 1)].alignment = self.align_cent
            sheet["D" + str(row + 1)].alignment = self.align_cent
            sheet["C" + str(row + 2)].alignment = self.align_cent
            sheet["D" + str(row + 2)].alignment = self.align_cent
            sheet["E" + str(row + 2)].alignment = self.align_cent
            sheet["F" + str(row + 2)].alignment = self.align_cent
            sheet["G" + str(row + 2)].alignment = self.align_cent
            sheet["B" + str(row + 3)].alignment = self.align_cent
            sheet["C" + str(row + 3)].alignment = self.align_cent
            sheet["D" + str(row + 3)].alignment = self.align_cent
            sheet["E" + str(row + 3)].alignment = self.align_cent
            sheet["F" + str(row + 3)].alignment = self.align_cent
            sheet["G" + str(row + 3)].alignment = self.align_cent


