import openpyxl.worksheet
from openpyxl.styles import PatternFill, Font

class PaintDesigner():
    
    def __init__(self) -> None:
        # define color
        self.bg_color_title = PatternFill(patternType="solid", fgColor="e6eeff")

        self.sheet = openpyxl.worksheet
        self.default_font = Font(size=8)

    def paint_individual(self, sheet: openpyxl.worksheet):
        self.sheet = sheet
        # fontサイズ調整
        self.cange_fontsizes("A1:O150", self.default_font)
        self.company_summary(sheet)
        self.ipo_info(sheet)
        self.listing_forecast(sheet)
        self.financial_info(sheet)
        self.news_list(sheet)
        self.lockup_info(sheet)
        pass
    
    def company_summary(self, sheet: openpyxl.worksheet):
        sheet['A1'].fill = self.bg_color_title
        sheet['C1'].fill = self.bg_color_title
        self.paint_merge_cells("A3:A15", self.bg_color_title)
        self.paint_merge_cells("A17:G17", self.bg_color_title)
        self.paint_merge_cells("A18:A20", self.bg_color_title)
        self.paint_merge_cells("D18:E20", self.bg_color_title)
        pass
    
    def ipo_info(self, sheet):
        self.paint_merge_cells("A22:G22", self.bg_color_title)
        self.paint_merge_cells("A23:B25", self.bg_color_title)
        self.paint_merge_cells("A26:A29", self.bg_color_title)
        self.paint_merge_cells("A30:B31", self.bg_color_title)
        sheet['E23'].fill = self.bg_color_title
        self.paint_merge_cells("D26:E26", self.bg_color_title)
        self.paint_merge_cells("F26:F29", self.bg_color_title)
        self.paint_merge_cells("D30:D31", self.bg_color_title)
        sheet['G30'].fill = self.bg_color_title
        pass
    
    def listing_forecast(self, sheet: openpyxl.worksheet):
        self.paint_merge_cells("A33:G34", self.bg_color_title)
        self.paint_merge_cells("A35:A39", self.bg_color_title)
        self.paint_merge_cells("E35:E38", self.bg_color_title)
    
    def financial_info(self, sheet: openpyxl.worksheet):
        self.paint_merge_cells("A41:G41", self.bg_color_title)
        self.paint_merge_cells("A42:A49", self.bg_color_title)
        self.paint_merge_cells("A51:G51", self.bg_color_title)
        self.paint_merge_cells("A52:A58", self.bg_color_title)

    def news_list(self, sheet: openpyxl.worksheet):
        self.paint_merge_cells("A60:G61", self.bg_color_title)
        
    def lockup_info(self, sheet: openpyxl.worksheet):
        cell_row = 82
        self.paint_merge_cells("A" + str(cell_row) + ":G" + str(cell_row), self.bg_color_title)
        
        for row in range(cell_row + 1, cell_row + 40, 4):
            self.paint_merge_cells("A"+ str(row) + ":" + "A" + str(row + 3), self.bg_color_title)
            self.paint_merge_cells("C"+ str(row) + ":" + "C" + str(row + 2), self.bg_color_title)
            self.paint_merge_cells("D"+ str(row+ 2) + ":" + "D" + str(row + 3), self.bg_color_title)
            self.paint_merge_cells("F"+ str(row+ 2) + ":" + "F" + str(row + 3), self.bg_color_title)
            sheet["B" + str(row)].fill = self.bg_color_title
            sheet["B" + str(row + 3)].fill = self.bg_color_title
        
        
    def paint_merge_cells(self, range: str, pattern_fill: PatternFill):
        cells = self.sheet[range]
        for row in cells:
            for cell in row:
                cell.fill = pattern_fill

    def cange_fontsizes(self, range: str, font: Font):
        cells = self.sheet[range]
        for row in cells:
            for cell in row:
                cell.font = font
